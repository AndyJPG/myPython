import openpyxl as xl
from pathlib import Path
from openpyxl.chart import BarChart, Reference


# Getting file safely
xl_file = 'transactions.xlsx'

try:
    path = Path(xl_file).resolve(strict=True)
except FileNotFoundError:
    print('File does not exist')

my_file = Path(xl_file)
if not my_file.is_file():
    print('File not exist')


# Process excel file
def process_workbook(filename):
    work_book = xl.load_workbook(filename)
    sheet1 = work_book['Sheet1']

    # Experiment getting cell
    cell = sheet1['a1']
    cell = sheet1.cell(1, 3)
    print(cell.value)

    for row in range(2, sheet1.max_row + 1):
        cell = sheet1.cell(row, 3)
        corrected_price = cell.value * 0.9

        corrected_price_cell = sheet1.cell(row, 4)
        corrected_price_cell.value = round(corrected_price, 2)
        print(cell.value)

    values = Reference(sheet1,
              min_row=2,
              max_row=sheet1.max_row,
              min_col=4,
              max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet1.add_chart(chart, 'e2')

    saved_filename = filename.split('.')
    work_book.save(f'{saved_filename[0]}2.{saved_filename[-1]}')

process_workbook(xl_file)

print(Path('.').iterdir())
for file in Path('.').iterdir():
    print(file)